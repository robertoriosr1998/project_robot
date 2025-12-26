from openpyxl import Workbook

# Create workbook
wb = Workbook()

# Parameters sheet (rename default sheet)
ws_params = wb.active
ws_params.title = 'Parameters'
ws_params['A1'] = 'OPC folder:'

# OPC sheet with Column 1-50 headers
ws_opc = wb.create_sheet('OPC')
for i in range(1, 51):
    ws_opc.cell(row=1, column=i, value=f'Column {i}')

# TIPS sheet with headers
ws_tips = wb.create_sheet('TIPS')
tips_headers = ['Nº', 'FUND HOUSE', 'DAY', 'TIME', 'CHANGED TO TRADITIONAL', 'CONNECTIVITY', 
                'TOLERANCE', 'EXCEPTO TOLERANCIA', 'MAIL EMAIL ADRESS', 'ADDITIONAL EMAIL ADRESS', 
                'TIPS', 'COMMENTS', 'QUERY EMAIL 1', 'QUERY EMAIL 2', 'QUERY EMAIL 3', 
                'PORTAL LINK', 'MY TIPS', 'DOCUMENT PASSWORD 1', 'DOCUMENT PASSWORD 2', 'DOCUMENT PASSWORD 3']
for i, header in enumerate(tips_headers, 1):
    ws_tips.cell(row=1, column=i, value=header)

# CN Database sheet with headers
ws_cn = wb.create_sheet('CN Database')
cn_headers = ['ID', 'File Path', 'Is it a CN?', 'Operation Type', 'Is it a Multiseries?', 
              'Currency', 'Gross Amount', 'Net Amount', 'Units', 'Equalization', 
              'NAV price', 'NAV date', 'Settlement Date']
for i, header in enumerate(cn_headers, 1):
    ws_cn.cell(row=1, column=i, value=header)

# Save as xlsx (valid Excel format)
wb.save('OPC_TEST.xlsx')
print('OPC_TEST.xlsx created successfully!')
print('Sheets: Parameters, OPC, TIPS, CN Database')
print('')
print('Note: Saved as .xlsx since your Excel license has expired.')
print('To add macros, you would need an active Excel license.')
